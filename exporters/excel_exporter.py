# -*- coding: utf-8 -*-
"""
Professional Excel GNSS Accuracy Report Exporter
Generates structured multi-sheet .xlsx reports with executive summary, epoch-by-epoch raw errors, and anomaly events.
Powered by openpyxl with corporate style formatting.
"""

import os
from datetime import datetime
from PySide6.QtCore import Qt
from PySide6.QtWidgets import QFileDialog, QMessageBox, QProgressDialog, QApplication

def export_excel_report(parent_window, segments, truth, table_metrics, config=None):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        QMessageBox.warning(parent_window, "提示", "未检测到 openpyxl 模块。请在终端运行: pip install openpyxl")
        return

    if not segments:
        QMessageBox.warning(parent_window, "提示", "没有测试数据可供导出。")
        return

    save_path, _ = QFileDialog.getSaveFileName(
        parent_window, "导出 Excel 精度分析报表", "GNSS定位精度分析报表.xlsx", "Excel Files (*.xlsx)"
    )
    if not save_path:
        return

    progress = QProgressDialog("正在生成 Excel 报表...", "取消", 0, 100, parent_window)
    progress.setWindowTitle("导出 Excel")
    progress.setWindowModality(Qt.WindowModal)
    progress.setMinimumDuration(0)
    progress.setValue(5)
    QApplication.processEvents()

    try:
        wb = openpyxl.Workbook()
        
        # 预设企业级科技蓝样式
        header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        title_font = Font(name="Microsoft YaHei", size=14, bold=True, color="0F172A")
        sub_font = Font(name="Microsoft YaHei", size=10, italic=True, color="64748B")
        cell_font = Font(name="Consolas", size=10)
        bold_cell_font = Font(name="Consolas", size=10, bold=True)
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        alert_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # 淡红警示

        # =========================================================================
        # Sheet 1: 测试概览与精度指标
        # =========================================================================
        ws1 = wb.active
        ws1.title = "测试概览与精度指标"
        ws1.views.sheetView[0].showGridLines = True

        ws1['A1'] = "GNSS 高精度定位测试与精度评定报告"
        ws1['A1'].font = title_font
        ws1['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  分析工具: POS_Handling"
        ws1['A2'].font = sub_font

        # 写入精度指标表格 (从 table_metrics 读取)
        start_row = 4
        col_count = table_metrics.columnCount() if table_metrics else 0
        row_count = table_metrics.rowCount() if table_metrics else 0

        if col_count > 0:
            for c in range(col_count):
                cell = ws1.cell(row=start_row, column=c+1)
                cell.value = table_metrics.horizontalHeaderItem(c).text() if table_metrics.horizontalHeaderItem(c) else f"Col{c}"
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            ws1.row_dimensions[start_row].height = 24

            for r in range(row_count):
                curr_row = start_row + 1 + r
                fill = zebra_fill if r % 2 == 1 else PatternFill(fill_type=None)
                ws1.row_dimensions[curr_row].height = 20
                for c in range(col_count):
                    item = table_metrics.item(r, c)
                    val_str = item.text() if item else ""
                    cell = ws1.cell(row=curr_row, column=c+1)
                    
                    # 尝试转换数值型
                    try:
                        if val_str.endswith('%'):
                            cell.value = float(val_str.rstrip('%')) / 100.0
                            cell.number_format = '0.00%'
                        elif val_str.endswith('m'):
                            cell.value = float(val_str.rstrip('m').strip())
                            cell.number_format = '0.000'
                        else:
                            cell.value = float(val_str)
                            cell.number_format = '0.000'
                    except Exception:
                        cell.value = val_str

                    cell.font = cell_font
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

        progress.setValue(30)
        QApplication.processEvents()

        # =========================================================================
        # Sheet 2: 逐历元原始误差明细 (Epoch-by-Epoch Raw Errors)
        # =========================================================================
        ws2 = wb.create_sheet(title="逐历元原始误差明细")
        ws2.views.sheetView[0].showGridLines = True

        raw_headers = [
            "历元序号", "分段名称", "时间 (TOW/UTC)", "纬度 (Lat, °)", "经度 (Lon, °)", "高程 (Alt, m)",
            "东向偏差 (dE, m)", "北向偏差 (dN, m)", "天向偏差 (dU, m)", "水平偏差 (2D, m)", "三维偏差 (3D, m)",
            "解状态 (Fix Type)", "在用卫星数", "差分龄期 (Age, s)"
        ]

        for c, h in enumerate(raw_headers, 1):
            cell = ws2.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws2.row_dimensions[1].height = 24

        curr_row = 2
        anomalies = [] # 收集异常与离群历元

        active_segs = [s for s in segments if s.get('active', True)]
        total_epochs = sum(len(s.get('epochs', [])) for s in active_segs)
        processed_epochs = 0

        for seg in active_segs:
            seg_name = seg.get('name', '未命名分段')
            epochs = seg.get('epochs', [])
            metrics = seg.get('metrics', {})
            de_list = metrics.get('de', metrics.get('de_list', []))
            dn_list = metrics.get('dn', metrics.get('dn_list', []))
            du_list = metrics.get('v_errors', metrics.get('du_list', []))
            h_err_list = metrics.get('h_errors', [])

            if not de_list and 'enu_points' in metrics and metrics['enu_points']:
                de_list = [p.get('e', 0.0) for p in metrics['enu_points']]
                dn_list = [p.get('n', 0.0) for p in metrics['enu_points']]
                du_list = [p.get('u', 0.0) for p in metrics['enu_points']]

            for idx, ep in enumerate(epochs):
                if progress.wasCanceled():
                    return

                de = de_list[idx] if idx < len(de_list) else None
                dn = dn_list[idx] if idx < len(dn_list) else None
                du = du_list[idx] if idx < len(du_list) else None
                h_err = h_err_list[idx] if idx < len(h_err_list) else None
                pos_3d = ((de**2 + dn**2 + du**2)**0.5) if (de is not None and dn is not None and du is not None) else None

                fix_type = ep.get('quality', ep.get('status', 1))
                sats = ep.get('num_sats', ep.get('satellites', 0))
                age = ep.get('diff_age', ep.get('age', ''))

                time_val = ep.get('time_str', ep.get('utc_seconds', idx+1))

                row_vals = [
                    idx + 1,
                    seg_name,
                    time_val,
                    ep.get('lat', 0.0),
                    ep.get('lon', 0.0),
                    ep.get('alt', 0.0),
                    de,
                    dn,
                    du,
                    h_err,
                    pos_3d,
                    fix_type,
                    sats,
                    age
                ]

                # 判定是否为异常点: 水平误差大于 0.2m (或非RTK固定解4)
                is_anomaly = False
                if fix_type not in [4, '4', 'FIX', 'RTK_FIX'] or (h_err is not None and h_err > 0.3):
                    is_anomaly = True
                    anomalies.append(row_vals)

                fill = zebra_fill if (curr_row % 2 == 1) else PatternFill(fill_type=None)
                if is_anomaly:
                    fill = alert_fill

                ws2.row_dimensions[curr_row].height = 19
                for col_idx, val in enumerate(row_vals, 1):
                    cell = ws2.cell(row=curr_row, column=col_idx)
                    cell.value = val
                    cell.font = cell_font
                    cell.fill = fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                    # 数字精度格式
                    if isinstance(val, float):
                        if col_idx in [4, 5]: # 经纬度
                            cell.number_format = '0.00000000'
                        elif col_idx in [6, 7, 8, 9, 10, 11]: # 误差
                            cell.number_format = '0.000'

                curr_row += 1
                processed_epochs += 1
                if processed_epochs % 500 == 0 and total_epochs > 0:
                    pct = 30 + int(processed_epochs / total_epochs * 45)
                    progress.setValue(min(pct, 75))
                    QApplication.processEvents()

        # =========================================================================
        # Sheet 3: 异常与离群历元分析
        # =========================================================================
        ws3 = wb.create_sheet(title="异常与离群历元清单")
        ws3.views.sheetView[0].showGridLines = True

        for c, h in enumerate(raw_headers, 1):
            cell = ws3.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws3.row_dimensions[1].height = 24

        for r_idx, a_row in enumerate(anomalies, 2):
            ws3.row_dimensions[r_idx].height = 19
            for c_idx, val in enumerate(a_row, 1):
                cell = ws3.cell(row=r_idx, column=c_idx, value=val)
                cell.font = cell_font
                cell.fill = alert_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(val, float):
                    if c_idx in [4, 5]:
                        cell.number_format = '0.00000000'
                    elif c_idx in [6, 7, 8, 9, 10, 11]:
                        cell.number_format = '0.000'

        progress.setValue(85)
        QApplication.processEvents()

        # 自适应调整所有 Sheet 列宽
        for ws in [ws1, ws2, ws3]:
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row == 1 and ws.title == "测试概览与精度指标":
                        continue
                    val_str = str(cell.value or '')
                    max_len = max(max_len, len(val_str.encode('gbk', errors='ignore')))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        progress.setValue(95)
        QApplication.processEvents()

        # 保存文件 (带文件占用异常保护)
        wb.save(save_path)
        progress.setValue(100)
        QMessageBox.information(parent_window, "成功", f"Excel 评定报表已成功导出至:\n{save_path}")

    except PermissionError:
        QMessageBox.critical(parent_window, "文件被占用", f"无法写入文件:\n{save_path}\n请先关闭已在 Excel 中打开的同名文件后再试！")
    except Exception as e:
        QMessageBox.critical(parent_window, "导出失败", f"生成 Excel 报告时发生错误:\n{e}")
